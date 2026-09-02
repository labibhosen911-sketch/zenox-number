import os
import io
import sqlite3
import asyncio
from datetime import datetime, timezone

from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from playwright.async_api import async_playwright

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    KeyboardButton,
)
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

# ============================================================
# ZENOX NUMBER BOT
# Safe version:
# - Number inventory/reservation
# - Admin dashboard
# - Panel management + inventory sync
# - XLSX import/export
# - Join requirements
# - Admin management
# - Announcements
# - Support
# - OTP group METADATA management only
#
# This script intentionally DOES NOT read, capture, display,
# or forward SMS/OTP/verification codes.
# ============================================================

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
DB_PATH = os.getenv("DB_PATH", "bot.db")
CHANNEL_ID = os.getenv("CHANNEL_ID", "").strip()
CHECK_SECONDS = int(os.getenv("CHECK_SECONDS", "15"))

SUPPORT_ID = os.getenv("SUPPORT_ID", "@labibhosen74").strip()
OWNER_ID = int(os.getenv("OWNER_ID", "0") or 0)
OWNER_USERNAME = os.getenv("OWNER_USERNAME", "labibhosen74").strip().lstrip("@").lower()

# Optional static admins. Owner can add/remove managed admins from the bot.
STATIC_ADMIN_IDS = {
    int(x.strip())
    for x in os.getenv("ADMIN_IDS", "").split(",")
    if x.strip().isdigit()
}

# Panel table selectors. Change these for your authorized panel if needed.
PANEL_USERNAME_SELECTOR = os.getenv("PANEL_USERNAME_SELECTOR", 'input[name="username"]')
PANEL_PASSWORD_SELECTOR = os.getenv("PANEL_PASSWORD_SELECTOR", 'input[name="password"]')
PANEL_LOGIN_SELECTOR = os.getenv("PANEL_LOGIN_SELECTOR", 'button[type="submit"]')
PANEL_TABLE_SELECTOR = os.getenv("PANEL_TABLE_SELECTOR", "table")
PANEL_HEADLESS = os.getenv("PANEL_HEADLESS", "true").lower() == "true"


# ============================================================
# DATABASE
# ============================================================

db = sqlite3.connect(DB_PATH, check_same_thread=False)
db.row_factory = sqlite3.Row
db_lock = asyncio.Lock()


def utc_now():
    return datetime.now(timezone.utc).isoformat()


def sql(query, params=(), fetch=False):
    cur = db.cursor()
    cur.execute(query, params)
    rows = cur.fetchall() if fetch else None
    db.commit()
    return rows if fetch else cur.lastrowid


def init_db():
    sql("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT DEFAULT '',
            first_name TEXT DEFAULT '',
            balance REAL DEFAULT 0,
            created_at TEXT
        )
    """)

    sql("""
        CREATE TABLE IF NOT EXISTS numbers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            number TEXT UNIQUE NOT NULL,
            country TEXT DEFAULT '',
            status TEXT DEFAULT 'available',
            reserved_by INTEGER,
            reserved_at TEXT,
            panel_status TEXT DEFAULT '',
            prefix TEXT DEFAULT '',
            otp_group TEXT DEFAULT '',
            source_panel TEXT DEFAULT ''
        )
    """)

    sql("""
        CREATE TABLE IF NOT EXISTS panels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            url TEXT NOT NULL,
            username TEXT NOT NULL,
            password TEXT NOT NULL,
            active INTEGER DEFAULT 1,
            last_error TEXT DEFAULT '',
            last_sync TEXT DEFAULT '',
            created_at TEXT
        )
    """)

    sql("""
        CREATE TABLE IF NOT EXISTS join_requirements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            channel TEXT UNIQUE NOT NULL,
            title TEXT DEFAULT '',
            url TEXT DEFAULT '',
            created_at TEXT
        )
    """)

    sql("""
        CREATE TABLE IF NOT EXISTS otp_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            chat_id TEXT UNIQUE NOT NULL,
            active INTEGER DEFAULT 1,
            created_at TEXT
        )
    """)

    sql("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT DEFAULT ''
        )
    """)

    for key, value in {
        "announcement": "",
        "support": SUPPORT_ID,
        "auto_delete": "off",
        "strict_one": "off",
        "managed_admins": "",
    }.items():
        sql(
            "INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)",
            (key, value),
        )


def setting(key):
    rows = sql("SELECT value FROM settings WHERE key=?", (key,), True)
    return rows[0]["value"] if rows else ""


def set_setting(key, value):
    sql(
        "INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",
        (key, value),
    )


def managed_admin_ids():
    return {
        int(x.strip())
        for x in setting("managed_admins").split(",")
        if x.strip().isdigit()
    }


def save_managed_admin_ids(ids):
    set_setting("managed_admins", ",".join(map(str, sorted(set(ids)))))


def is_owner(user):
    uid = getattr(user, "id", 0)
    username = (getattr(user, "username", "") or "").lstrip("@").lower()
    return (OWNER_ID and uid == OWNER_ID) or (
        OWNER_USERNAME and username == OWNER_USERNAME
    )


def is_admin(user_or_id):
    uid = user_or_id if isinstance(user_or_id, int) else getattr(user_or_id, "id", 0)
    return uid in STATIC_ADMIN_IDS or uid in managed_admin_ids() or (
        not isinstance(user_or_id, int) and is_owner(user_or_id)
    )


def register_user(user):
    sql("""
        INSERT OR IGNORE INTO users
        (user_id,username,first_name,created_at)
        VALUES(?,?,?,?)
    """, (user.id, user.username or "", user.first_name or "", utc_now()))

    sql("""
        UPDATE users SET username=?, first_name=? WHERE user_id=?
    """, (user.username or "", user.first_name or "", user.id))


def add_number(number, country="", panel_status="", prefix="",
               otp_group="", source_panel=""):
    number = str(number).strip()
    if not number:
        return False

    existing = sql(
        "SELECT id FROM numbers WHERE number=?",
        (number,),
        True,
    )
    if existing:
        sql("""
            UPDATE numbers
            SET country=?, panel_status=?, prefix=?, otp_group=?, source_panel=?
            WHERE number=?
        """, (country, panel_status, prefix, otp_group, source_panel, number))
        return False

    sql("""
        INSERT INTO numbers
        (number,country,status,panel_status,prefix,otp_group,source_panel)
        VALUES(?,?,?,?,?,?,?)
    """, (
        number, country, "available", panel_status,
        prefix, otp_group, source_panel,
    ))
    return True


def available_numbers():
    return sql("""
        SELECT * FROM numbers
        WHERE status='available' AND reserved_by IS NULL
        ORDER BY id
    """, (), True)


def get_number(number_id):
    rows = sql("SELECT * FROM numbers WHERE id=?", (number_id,), True)
    return rows[0] if rows else None


def reserve_number(number_id, user_id):
    cur = db.cursor()
    cur.execute("""
        UPDATE numbers
        SET status='reserved', reserved_by=?, reserved_at=?
        WHERE id=? AND status='available' AND reserved_by IS NULL
    """, (user_id, utc_now(), number_id))
    db.commit()
    return cur.rowcount == 1


def panels():
    return sql("SELECT * FROM panels ORDER BY id", (), True)


def add_panel(name, url, username, password):
    sql("""
        INSERT INTO panels(name,url,username,password,created_at)
        VALUES(?,?,?,?,?)
    """, (name, url, username, password, utc_now()))


def toggle_panel(panel_id):
    sql("""
        UPDATE panels
        SET active=CASE WHEN active=1 THEN 0 ELSE 1 END
        WHERE id=?
    """, (panel_id,))


def delete_panel(panel_id):
    sql("DELETE FROM panels WHERE id=?", (panel_id,))


def join_requirements():
    return sql("SELECT * FROM join_requirements ORDER BY id", (), True)


def add_join(channel, title="", url=""):
    try:
        sql("""
            INSERT INTO join_requirements(channel,title,url,created_at)
            VALUES(?,?,?,?)
        """, (channel, title, url, utc_now()))
        return True
    except sqlite3.IntegrityError:
        return False


def delete_join(row_id):
    sql("DELETE FROM join_requirements WHERE id=?", (row_id,))


def otp_groups():
    return sql("SELECT * FROM otp_groups ORDER BY id", (), True)


def add_otp_group(name, chat_id):
    try:
        sql("""
            INSERT INTO otp_groups(name,chat_id,created_at)
            VALUES(?,?,?)
        """, (name, chat_id, utc_now()))
        return True
    except sqlite3.IntegrityError:
        return False


def toggle_otp_group(row_id):
    sql("""
        UPDATE otp_groups
        SET active=CASE WHEN active=1 THEN 0 ELSE 1 END
        WHERE id=?
    """, (row_id,))


def delete_otp_group(row_id):
    sql("DELETE FROM otp_groups WHERE id=?", (row_id,))


# ============================================================
# UI
# ============================================================

BTN_GET = "💎 Get Number"
BTN_WALLET = "💰 Wallet"
BTN_STATUS = "📊 Status"
BTN_ADMIN = "🛡️ Admin Panel"
BTN_SUPPORT = "🆘 Support"


def support_value():
    return setting("support") or SUPPORT_ID


def support_url():
    value = support_value().strip()
    value = value.replace("https://t.me/", "").replace("http://t.me/", "")
    value = value.replace("t.me/", "").lstrip("@").split("/")[0]
    return f"https://t.me/{value}"


def main_inline(user):
    rows = [
        [InlineKeyboardButton("💎 𝗚𝗘𝗧 𝗡𝗨𝗠𝗕𝗘𝗥", callback_data="get_numbers")],
        [
            InlineKeyboardButton("💰 𝗪𝗔𝗟𝗟𝗘𝗧", callback_data="wallet"),
            InlineKeyboardButton("📊 𝗦𝗧𝗔𝗧𝗨𝗦", callback_data="status"),
        ],
    ]
    if is_admin(user):
        rows.append([
            InlineKeyboardButton("🛡️ 𝗔𝗗𝗠𝗜𝗡 𝗣𝗔𝗡𝗘𝗟", callback_data="admin")
        ])
    else:
        rows.append([
            InlineKeyboardButton("🆘 𝗦𝗨𝗣𝗣𝗢𝗥𝗧", callback_data="support")
        ])
    return InlineKeyboardMarkup(rows)


def reply_menu(user):
    rows = [
        [KeyboardButton(BTN_GET)],
        [KeyboardButton(BTN_WALLET), KeyboardButton(BTN_STATUS)],
    ]
    rows.append([
        KeyboardButton(BTN_ADMIN if is_admin(user) else BTN_SUPPORT)
    ])
    return ReplyKeyboardMarkup(
        rows,
        resize_keyboard=True,
        is_persistent=True,
        input_field_placeholder="Select an option...",
    )


def back_button(target="back"):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("↩️ 𝗕𝗔𝗖𝗞", callback_data=target)]
    ])


def admin_menu():
    auto = setting("auto_delete") == "on"
    strict = setting("strict_one") == "on"

    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 𝗦𝗧𝗔𝗧𝗜𝗦𝗧𝗜𝗖𝗦", callback_data="statistics")],
        [InlineKeyboardButton("📥 𝗨𝗣𝗟𝗢𝗔𝗗 𝗡𝗨𝗠𝗕𝗘𝗥𝗦", callback_data="upload")],
        [
            InlineKeyboardButton("🗑️ 𝗗𝗘𝗟𝗘𝗧𝗘 𝗥𝗔𝗡𝗚𝗘", callback_data="delete"),
            InlineKeyboardButton("📤 𝗘𝗫𝗣𝗢𝗥𝗧", callback_data="export"),
        ],
        [InlineKeyboardButton("🛠️ 𝗪𝗢𝗥𝗞𝗘𝗥 𝗠𝗔𝗡𝗔𝗚𝗘𝗠𝗘𝗡𝗧", callback_data="workers")],
        [InlineKeyboardButton("🔐 𝗝𝗢𝗜𝗡 𝗥𝗘𝗤𝗨𝗜𝗥𝗘𝗠𝗘𝗡𝗧𝗦", callback_data="join")],
        [InlineKeyboardButton("🛡️ 𝗢𝗧𝗣 𝗚𝗥𝗢𝗨𝗣𝗦", callback_data="otp_groups")],
        [InlineKeyboardButton("📣 𝗔𝗡𝗡𝗢𝗨𝗡𝗖𝗘𝗠𝗘𝗡𝗧", callback_data="announcement")],
        [
            InlineKeyboardButton(
                f"🧹 Auto-Delete: {'🟢 ON' if auto else '🔴 OFF'}",
                callback_data="toggle_auto",
            ),
            InlineKeyboardButton(
                f"🛡️ Strict 1: {'🟢 ON' if strict else '🔴 OFF'}",
                callback_data="toggle_strict",
            ),
        ],
        [InlineKeyboardButton("👥 𝗠𝗔𝗡𝗔𝗚𝗘 𝗔𝗗𝗠𝗜𝗡𝗦", callback_data="admins")],
        [InlineKeyboardButton("✈️ 𝗦𝗘𝗧 𝗦𝗨𝗣𝗣𝗢𝗥𝗧 𝗜𝗗", callback_data="set_support")],
        [InlineKeyboardButton("↩️ 𝗕𝗔𝗖𝗞", callback_data="back")],
    ])


async def edit_or_send(target, text, markup):
    if hasattr(target, "edit_message_text"):
        await target.edit_message_text(
            text, parse_mode=ParseMode.HTML, reply_markup=markup
        )
    else:
        await target.reply_text(
            text, parse_mode=ParseMode.HTML, reply_markup=markup
        )


# ============================================================
# JOIN REQUIREMENT
# ============================================================

async def missing_joined(bot, user_id):
    missing = []
    for row in join_requirements():
        try:
            member = await bot.get_chat_member(row["channel"], user_id)
            if member.status not in {"member", "administrator", "creator"}:
                missing.append(row)
        except Exception:
            missing.append(row)
    return missing


def join_markup(rows):
    buttons = []
    for row in rows:
        url = row["url"]
        if not url and str(row["channel"]).startswith("@"):
            url = f"https://t.me/{str(row['channel']).lstrip('@')}"
        if url:
            buttons.append([
                InlineKeyboardButton(
                    f"📣 Join {row['title'] or row['channel']}",
                    url=url,
                )
            ])
    buttons.append([
        InlineKeyboardButton("✅ I Joined — Check", callback_data="check_join")
    ])
    return InlineKeyboardMarkup(buttons)


async def ensure_joined(target, context):
    user = target.from_user
    if is_admin(user):
        return True
    missing = await missing_joined(context.bot, user.id)
    if missing:
        await edit_or_send(
            target,
            "🔐 <b>JOIN REQUIREMENT</b>\n\n"
            "Please join every required channel, then press "
            "<b>✅ I Joined — Check</b>.",
            join_markup(missing),
        )
        return False
    return True


# ============================================================
# USER PAGES
# ============================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    register_user(user)

    missing = await missing_joined(context.bot, user.id) if not is_admin(user) else []
    if missing:
        await update.message.reply_text(
            "🔐 <b>JOIN REQUIREMENT</b>\n\n"
            "Join the required channel(s) and press the check button.",
            parse_mode=ParseMode.HTML,
            reply_markup=join_markup(missing),
        )
        return

    announcement = setting("announcement")
    text = (
        "💎 <b>𝗪𝗘𝗟𝗖𝗢𝗠𝗘 𝗧𝗢 𝗭𝗘𝗡𝗢𝗫</b> 💎\n\n"
        "📲 <b>Number Service</b>\n\n"
        "✨ Choose an option below."
    )
    if announcement:
        text += f"\n\n📣 <b>Announcement</b>\n{announcement}"

    await update.message.reply_text(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=reply_menu(user),
    )


async def show_numbers(target):
    rows = available_numbers()
    if not rows:
        await edit_or_send(
            target,
            "❌ <b>NO NUMBERS AVAILABLE</b>\n\n⏳ Please try again later.",
            back_button(),
        )
        return

    buttons = []
    for row in rows[:40]:
        buttons.append([
            InlineKeyboardButton(
                f"📱 {row['number']}",
                callback_data=f"number:{row['id']}",
            )
        ])
    buttons.append([
        InlineKeyboardButton("↩️ 𝗕𝗔𝗖𝗞", callback_data="back")
    ])

    await edit_or_send(
        target,
        "💎 <b>𝗔𝗩𝗔𝗜𝗟𝗔𝗕𝗟𝗘 𝗡𝗨𝗠𝗕𝗘𝗥𝗦</b>\n\n"
        "Select a number:",
        InlineKeyboardMarkup(buttons),
    )


async def number_detail(query, row_id):
    row = get_number(row_id)
    if not row:
        await query.answer("Number no longer exists.", show_alert=True)
        return

    prefix = row["prefix"] or "OFF"
    group = row["otp_group"] or "Not assigned"

    text = (
        "📱 <b>𝗡𝗨𝗠𝗕𝗘𝗥</b>\n\n"
        f"🔢 <b>Number:</b> <code>{row['number']}</code>\n"
        f"🏷 <b>Prefix:</b> <code>{prefix}</code>\n"
        f"🛡️ <b>OTP Group:</b> <code>{group}</code>\n"
        f"🌍 <b>Country:</b> {row['country'] or 'Unknown'}\n"
        f"📡 <b>Panel Status:</b> {row['panel_status'] or 'Available'}\n\n"
        "Choose an action:"
    )

    await query.edit_message_text(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ 𝗥𝗘𝗦𝗘𝗥𝗩𝗘 𝗡𝗨𝗠𝗕𝗘𝗥", callback_data=f"reserve:{row_id}")],
            [
                InlineKeyboardButton(f"🏷 Prefix: {prefix}", callback_data="noop"),
                InlineKeyboardButton(f"🛡 Group: {group}", callback_data="noop"),
            ],
            [InlineKeyboardButton("🔄 𝗖𝗛𝗔𝗡𝗚𝗘 𝗡𝗨𝗠𝗕𝗘𝗥", callback_data="get_numbers")],
            [InlineKeyboardButton("↩️ 𝗕𝗔𝗖𝗞", callback_data="get_numbers")],
        ]),
    )


async def reserve(query, row_id):
    user_id = query.from_user.id
    if not reserve_number(row_id, user_id):
        await query.answer("❌ This number is already reserved.", show_alert=True)
        return

    row = get_number(row_id)
    text = (
        "✅ <b>𝗡𝗨𝗠𝗕𝗘𝗥 𝗥𝗘𝗦𝗘𝗥𝗩𝗘𝗗</b>\n\n"
        f"📱 <b>Number:</b> <code>{row['number']}</code>\n"
        f"🏷 <b>Prefix:</b> <code>{row['prefix'] or 'OFF'}</code>\n"
        f"🛡️ <b>OTP Group:</b> <code>{row['otp_group'] or 'Not assigned'}</code>\n"
        f"🌍 <b>Country:</b> {row['country'] or 'Unknown'}\n\n"
        "🟢 Status: <b>Reserved</b>"
    )
    await query.edit_message_text(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=back_button(),
    )


async def wallet(target):
    rows = sql(
        "SELECT balance FROM users WHERE user_id=?",
        (target.from_user.id,),
        True,
    )
    balance = rows[0]["balance"] if rows else 0
    await edit_or_send(
        target,
        f"👛 <b>𝗬𝗢𝗨𝗥 𝗪𝗔𝗟𝗟𝗘𝗧</b>\n\n💰 Balance: <b>${balance:.2f}</b>",
        back_button(),
    )


async def status(target):
    users = sql("SELECT COUNT(*) c FROM users", (), True)[0]["c"]
    total = sql("SELECT COUNT(*) c FROM numbers", (), True)[0]["c"]
    available = sql(
        "SELECT COUNT(*) c FROM numbers WHERE status='available'",
        (), True
    )[0]["c"]
    reserved = sql(
        "SELECT COUNT(*) c FROM numbers WHERE status='reserved'",
        (), True
    )[0]["c"]

    await edit_or_send(
        target,
        "📊 <b>𝗦𝗬𝗦𝗧𝗘𝗠 𝗦𝗧𝗔𝗧𝗨𝗦</b>\n\n"
        f"👥 Users: <b>{users}</b>\n"
        f"📱 Total Numbers: <b>{total}</b>\n"
        f"🟢 Available: <b>{available}</b>\n"
        f"🔴 Reserved: <b>{reserved}</b>",
        back_button(),
    )


async def support_page(target):
    await edit_or_send(
        target,
        "🆘 <b>𝗔𝗗𝗠𝗜𝗡 𝗦𝗨𝗣𝗣𝗢𝗥𝗧</b>\n\n"
        f"👤 Support: <code>{support_value()}</code>\n\n"
        "Tap the button below to contact support.",
        InlineKeyboardMarkup([
            [InlineKeyboardButton("✈️ 𝗖𝗢𝗡𝗧𝗔𝗖𝗧 𝗦𝗨𝗣𝗣𝗢𝗥𝗧", url=support_url())],
            [InlineKeyboardButton("↩️ 𝗕𝗔𝗖𝗞", callback_data="back")],
        ]),
    )


# ============================================================
# ADMIN DASHBOARD
# ============================================================

async def admin_page(query):
    if not is_admin(query.from_user):
        await support_page(query)
        return

    await query.edit_message_text(
        "🛡️ <b>𝗔𝗗𝗠𝗜𝗡 𝗗𝗔𝗦𝗛𝗕𝗢𝗔𝗥𝗗</b>\n\n"
        "💎 Select an option:",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_menu(),
    )


async def statistics(query):
    if not is_admin(query.from_user):
        return
    users = sql("SELECT COUNT(*) c FROM users", (), True)[0]["c"]
    total = sql("SELECT COUNT(*) c FROM numbers", (), True)[0]["c"]
    available = sql(
        "SELECT COUNT(*) c FROM numbers WHERE status='available'", (), True
    )[0]["c"]
    reserved = sql(
        "SELECT COUNT(*) c FROM numbers WHERE status='reserved'", (), True
    )[0]["c"]
    await query.edit_message_text(
        "📊 <b>𝗦𝗧𝗔𝗧𝗜𝗦𝗧𝗜𝗖𝗦</b>\n\n"
        f"👥 Users: <b>{users}</b>\n"
        f"📱 Numbers: <b>{total}</b>\n"
        f"🟢 Available: <b>{available}</b>\n"
        f"🔴 Reserved: <b>{reserved}</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=back_button("admin"),
    )


# ============================================================
# UPLOAD / EXPORT / DELETE
# ============================================================

async def upload_start(query, context):
    if not is_admin(query.from_user):
        return
    context.user_data["mode"] = "upload"
    await query.edit_message_text(
        "📥 <b>𝗨𝗣𝗟𝗢𝗔𝗗 𝗡𝗨𝗠𝗕𝗘𝗥𝗦</b>\n\n"
        "Send an XLSX file.\n\n"
        "Recommended columns:\n"
        "<code>number | country | prefix | otp_group | panel</code>\n\n"
        "You can also send plain text, one number per line.",
        parse_mode=ParseMode.HTML,
        reply_markup=back_button("admin"),
    )


async def process_text_upload(update, context):
    added = 0
    for raw in (update.message.text or "").splitlines():
        parts = [x.strip() for x in raw.split("|")]
        number = parts[0] if parts else ""
        if not number:
            continue
        if add_number(
            number,
            parts[1] if len(parts) > 1 else "",
            prefix=parts[2] if len(parts) > 2 else "",
            otp_group=parts[3] if len(parts) > 3 else "",
            source_panel=parts[4] if len(parts) > 4 else "",
        ):
            added += 1

    context.user_data.pop("mode", None)
    await update.message.reply_text(
        f"✅ <b>{added}</b> new number(s) added.",
        parse_mode=ParseMode.HTML,
        reply_markup=reply_menu(update.effective_user),
    )


async def document_handler(update, context):
    if context.user_data.get("mode") != "upload" or not is_admin(update.effective_user):
        return

    doc = update.message.document
    if not doc.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("❌ Please upload an .xlsx file.")
        return

    file = await doc.get_file()
    payload = await file.download_as_bytearray()

    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)

    if not headers:
        wb.close()
        await update.message.reply_text("❌ XLSX is empty.")
        return

    normalized = [
        str(v or "").strip().lower().replace(" ", "_")
        for v in headers
    ]
    index = {name: i for i, name in enumerate(normalized)}

    def cell(row, *names):
        for name in names:
            if name in index and index[name] < len(row):
                return str(row[index[name]] or "").strip()
        return ""

    added = 0
    updated = 0

    for row in rows:
        number = cell(row, "number", "phone", "phone_number")
        if not number:
            continue

        was_added = add_number(
            number,
            cell(row, "country", "country_code"),
            cell(row, "status", "panel_status"),
            cell(row, "prefix"),
            cell(row, "otp_group", "group"),
            cell(row, "panel", "source_panel"),
        )
        if was_added:
            added += 1
        else:
            updated += 1

    wb.close()
    context.user_data.pop("mode", None)

    await update.message.reply_text(
        "✅ <b>XLSX import complete</b>\n\n"
        f"📥 Added: <b>{added}</b>\n"
        f"♻️ Updated: <b>{updated}</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=reply_menu(update.effective_user),
    )


async def export_numbers(query):
    if not is_admin(query.from_user):
        return

    rows = sql("""
        SELECT number,country,status,prefix,otp_group,source_panel,
               reserved_by,reserved_at
        FROM numbers ORDER BY id
    """, (), True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Numbers"
    ws.append([
        "number", "country", "status", "prefix", "otp_group",
        "panel", "reserved_by", "reserved_at"
    ])

    for row in rows:
        ws.append([
            row["number"], row["country"], row["status"], row["prefix"],
            row["otp_group"], row["source_panel"],
            row["reserved_by"] or "", row["reserved_at"] or "",
        ])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    await query.message.reply_document(
        document=out,
        filename="numbers.xlsx",
        caption="📤 XLSX export complete.",
    )


async def delete_start(query, context):
    if not is_admin(query.from_user):
        return
    context.user_data["mode"] = "delete"
    await query.edit_message_text(
        "🗑️ <b>𝗗𝗘𝗟𝗘𝗧𝗘 𝗥𝗔𝗡𝗚𝗘</b>\n\n"
        "Send ID range.\nExample: <code>1-20</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=back_button("admin"),
    )


async def process_delete(update, context):
    try:
        start, end = map(int, (update.message.text or "").strip().split("-"))
        cur = db.cursor()
        cur.execute(
            "DELETE FROM numbers WHERE id BETWEEN ? AND ?",
            (start, end),
        )
        db.commit()
        deleted = cur.rowcount
        text = f"✅ Deleted <b>{deleted}</b> number(s) from IDs <b>{start}-{end}</b>."
    except Exception:
        text = "❌ Invalid range. Use: <code>1-20</code>"

    context.user_data.pop("mode", None)
    await update.message.reply_text(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=reply_menu(update.effective_user),
    )


# ============================================================
# WORKER / PANEL MANAGEMENT
# ============================================================

async def workers_page(query):
    if not is_admin(query.from_user):
        return

    rows = panels()
    lines = [
        "🛠️ <b>𝗪𝗢𝗥𝗞𝗘𝗥 𝗠𝗔𝗡𝗔𝗚𝗘𝗠𝗘𝗡𝗧</b>",
        "",
        "Add your authorized panel. The bot will log in and sync its number inventory.",
        "",
    ]
    buttons = [
        [InlineKeyboardButton("➕ 𝗔𝗗𝗗 𝗡𝗘𝗪 𝗣𝗔𝗡𝗘𝗟", callback_data="add_panel")]
    ]

    if not rows:
        lines.append("No panels configured.")
    else:
        for row in rows:
            state = "🟢 ON" if row["active"] else "🔴 OFF"
            lines.append(f"• <b>{row['name']}</b> — {state}")
            if row["last_error"]:
                lines.append(f"  ⚠️ {row['last_error'][:120]}")
            buttons.append([
                InlineKeyboardButton(
                    f"{'🔴 Disable' if row['active'] else '🟢 Enable'}",
                    callback_data=f"panel_toggle:{row['id']}",
                ),
                InlineKeyboardButton(
                    "🗑️ Delete",
                    callback_data=f"panel_delete:{row['id']}",
                ),
            ])

    buttons.append([InlineKeyboardButton("↩️ 𝗔𝗗𝗠𝗜𝗡", callback_data="admin")])

    await query.edit_message_text(
        "\n".join(lines),
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def add_panel_start(query, context):
    if not is_admin(query.from_user):
        return
    context.user_data["mode"] = "panel_setup"
    context.user_data["panel"] = {}
    await query.edit_message_text(
        "🛠️ <b>𝗔𝗗𝗗 𝗣𝗔𝗡𝗘𝗟</b>\n\n"
        "Step 1/4 — send panel <b>name</b>.\nExample: <code>Panel 1</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=back_button("workers"),
    )


async def process_panel_setup(update, context):
    if not is_admin(update.effective_user):
        return

    value = (update.message.text or "").strip()
    if not value:
        await update.message.reply_text("❌ Please send a value.")
        return

    data = context.user_data.setdefault("panel", {})
    step = len(data)

    if step == 0:
        data["name"] = value
        prompt = "Step 2/4 — send panel <b>URL</b>.\nExample: <code>https://panel.example.com</code>"
    elif step == 1:
        if not value.startswith(("http://", "https://")):
            await update.message.reply_text("❌ URL must start with http:// or https://.")
            return
        data["url"] = value
        prompt = "Step 3/4 — send panel <b>username</b>."
    elif step == 2:
        data["username"] = value
        prompt = "Step 4/4 — send panel <b>password</b>."
    else:
        data["password"] = value
        add_panel(data["name"], data["url"], data["username"], data["password"])
        context.user_data.pop("mode", None)
        context.user_data.pop("panel", None)
        await update.message.reply_text(
            "✅ <b>Panel added.</b>\n\n"
            "🟢 Automatic inventory sync has been queued.\n"
            "The bot will only sync authorized number-inventory data; "
            "it does not capture or forward OTP/SMS codes.",
            parse_mode=ParseMode.HTML,
            reply_markup=reply_menu(update.effective_user),
        )
        return

    await update.message.reply_text(prompt, parse_mode=ParseMode.HTML)


# ============================================================
# JOIN REQUIREMENTS
# ============================================================

async def join_page(query):
    if not is_admin(query.from_user):
        return

    rows = join_requirements()
    lines = [
        "🔐 <b>𝗝𝗢𝗜𝗡 𝗥𝗘𝗤𝗨𝗜𝗥𝗘𝗠𝗘𝗡𝗧𝗦</b>",
        "",
        "Members must join all configured channels before using the bot.",
        "",
    ]
    buttons = [
        [InlineKeyboardButton("➕ 𝗔𝗗𝗗 𝗖𝗛𝗔𝗡𝗡𝗘𝗟", callback_data="join_add")]
    ]

    for row in rows:
        label = row["title"] or row["channel"]
        lines.append(f"• <b>{label}</b> — <code>{row['channel']}</code>")
        buttons.append([
            InlineKeyboardButton(
                f"🗑️ Remove {label}",
                callback_data=f"join_delete:{row['id']}",
            )
        ])

    buttons.append([InlineKeyboardButton("↩️ 𝗔𝗗𝗠𝗜𝗡", callback_data="admin")])

    await query.edit_message_text(
        "\n".join(lines),
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def join_add_start(query, context):
    if not is_admin(query.from_user):
        return
    context.user_data["mode"] = "join_add"
    await query.edit_message_text(
        "➕ <b>𝗔𝗗𝗗 𝗝𝗢𝗜𝗡 𝗥𝗘𝗤𝗨𝗜𝗥𝗘𝗠𝗘𝗡𝗧</b>\n\n"
        "Send:\n<code>@channel | Channel title | https://t.me/channel</code>\n\n"
        "The link is optional.",
        parse_mode=ParseMode.HTML,
        reply_markup=back_button("join"),
    )


async def process_join_add(update, context):
    parts = [x.strip() for x in (update.message.text or "").split("|")]
    channel = parts[0] if parts else ""
    title = parts[1] if len(parts) > 1 else channel
    url = parts[2] if len(parts) > 2 else ""

    if channel and not channel.startswith(("@", "-100")):
        channel = "@" + channel

    if not channel:
        await update.message.reply_text("❌ Invalid channel.")
        return

    ok = add_join(channel, title, url)
    context.user_data.pop("mode", None)
    await update.message.reply_text(
        "✅ Join requirement added." if ok else "⚠️ Already configured.",
        reply_markup=reply_menu(update.effective_user),
    )


# ============================================================
# OTP GROUP METADATA
# ============================================================

async def otp_groups_page(query):
    if not is_admin(query.from_user):
        return

    rows = otp_groups()
    lines = [
        "🛡️ <b>𝗢𝗧𝗣 𝗚𝗥𝗢𝗨𝗣𝗦</b>",
        "",
        "This section manages group destinations/metadata only.",
        "OTP/SMS contents are intentionally not captured or forwarded.",
        "",
    ]
    buttons = [
        [InlineKeyboardButton("➕ 𝗔𝗗𝗗 𝗚𝗥𝗢𝗨𝗣", callback_data="otp_add")]
    ]

    for row in rows:
        state = "🟢 ON" if row["active"] else "🔴 OFF"
        lines.append(f"• <b>{row['name']}</b> — <code>{row['chat_id']}</code> — {state}")
        buttons.append([
            InlineKeyboardButton(
                "🔄 Toggle",
                callback_data=f"otp_toggle:{row['id']}",
            ),
            InlineKeyboardButton(
                "🗑️ Delete",
                callback_data=f"otp_delete:{row['id']}",
            ),
        ])

    buttons.append([InlineKeyboardButton("↩️ 𝗔𝗗𝗠𝗜𝗡", callback_data="admin")])

    await query.edit_message_text(
        "\n".join(lines),
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def otp_add_start(query, context):
    if not is_admin(query.from_user):
        return
    context.user_data["mode"] = "otp_add"
    await query.edit_message_text(
        "➕ <b>𝗔𝗗𝗗 𝗢𝗧𝗣 𝗚𝗥𝗢𝗨𝗣</b>\n\n"
        "Send:\n<code>Group Name | -1001234567890</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=back_button("otp_groups"),
    )


async def process_otp_add(update, context):
    parts = [x.strip() for x in (update.message.text or "").split("|")]
    name = parts[0] if parts else ""
    chat_id = parts[1] if len(parts) > 1 else ""

    if not name or not chat_id:
        await update.message.reply_text(
            "❌ Use: <code>Group Name | -1001234567890</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    ok = add_otp_group(name, chat_id)
    context.user_data.pop("mode", None)
    await update.message.reply_text(
        "✅ OTP group added." if ok else "⚠️ Group already exists.",
        reply_markup=reply_menu(update.effective_user),
    )


# ============================================================
# ANNOUNCEMENT
# ============================================================

async def announcement_start(query, context):
    if not is_admin(query.from_user):
        return
    context.user_data["mode"] = "announcement"
    await query.edit_message_text(
        "📣 <b>𝗔𝗡𝗡𝗢𝗨𝗡𝗖𝗘𝗠𝗘𝗡𝗧</b>\n\n"
        "Send the announcement text. It will be sent to registered users.",
        parse_mode=ParseMode.HTML,
        reply_markup=back_button("admin"),
    )


async def process_announcement(update, context):
    if not is_admin(update.effective_user):
        return

    text = (update.message.text or "").strip()
    if not text:
        await update.message.reply_text("❌ Announcement cannot be empty.")
        return

    set_setting("announcement", text)
    sent = failed = 0

    for row in sql("SELECT user_id FROM users", (), True):
        try:
            await context.bot.send_message(
                row["user_id"],
                f"📣 <b>ANNOUNCEMENT</b>\n\n{text}",
                parse_mode=ParseMode.HTML,
            )
            sent += 1
            await asyncio.sleep(0.05)
        except Exception:
            failed += 1

    context.user_data.pop("mode", None)
    await update.message.reply_text(
        f"✅ Announcement sent.\n\n📨 Delivered: <b>{sent}</b>\n⚠️ Failed: <b>{failed}</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=reply_menu(update.effective_user),
    )


# ============================================================
# ADMIN MANAGEMENT — OWNER ONLY
# ============================================================

async def admins_page(query):
    if not is_owner(query.from_user):
        await query.answer("Only the owner can manage admins.", show_alert=True)
        return

    ids = sorted(managed_admin_ids())
    owner_label = f"@{OWNER_USERNAME}" if OWNER_USERNAME else str(OWNER_ID)

    lines = [
        "👥 <b>𝗠𝗔𝗡𝗔𝗚𝗘 𝗔𝗗𝗠𝗜𝗡𝗦</b>",
        "",
        f"👑 Owner: <code>{owner_label}</code>",
        "",
    ]
    buttons = [
        [InlineKeyboardButton("➕ 𝗔𝗗𝗗 𝗔𝗗𝗠𝗜𝗡", callback_data="admin_add")]
    ]

    if ids:
        for uid in ids:
            lines.append(f"• <code>{uid}</code>")
            buttons.append([
                InlineKeyboardButton(
                    f"🗑️ Remove {uid}",
                    callback_data=f"admin_remove:{uid}",
                )
            ])
    else:
        lines.append("No additional managed admins.")

    buttons.append([InlineKeyboardButton("↩️ 𝗔𝗗𝗠𝗜𝗡", callback_data="admin")])

    await query.edit_message_text(
        "\n".join(lines),
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def admin_add_start(query, context):
    if not is_owner(query.from_user):
        await query.answer("Owner only.", show_alert=True)
        return
    context.user_data["mode"] = "admin_add"
    await query.edit_message_text(
        "➕ <b>𝗔𝗗𝗗 𝗔𝗗𝗠𝗜𝗡</b>\n\nSend the numeric Telegram user ID.",
        parse_mode=ParseMode.HTML,
        reply_markup=back_button("admins"),
    )


async def process_admin_add(update, context):
    if not is_owner(update.effective_user):
        return

    value = (update.message.text or "").strip()
    if not value.isdigit():
        await update.message.reply_text("❌ User ID must be numeric.")
        return

    ids = managed_admin_ids()
    ids.add(int(value))
    save_managed_admin_ids(ids)
    context.user_data.pop("mode", None)

    await update.message.reply_text(
        f"✅ Admin <code>{value}</code> added.",
        parse_mode=ParseMode.HTML,
        reply_markup=reply_menu(update.effective_user),
    )


# ============================================================
# SUPPORT ID
# ============================================================

async def support_set_start(query, context):
    if not is_admin(query.from_user):
        return
    context.user_data["mode"] = "support_set"
    await query.edit_message_text(
        "✈️ <b>𝗦𝗘𝗧 𝗦𝗨𝗣𝗣𝗢𝗥𝗧 𝗜𝗗</b>\n\n"
        "Send username or t.me link.\nExample: <code>t.me/labibhosen74</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=back_button("admin"),
    )


async def process_support_set(update, context):
    if not is_admin(update.effective_user):
        return

    value = (update.message.text or "").strip()
    value = value.replace("https://t.me/", "").replace("http://t.me/", "")
    value = value.replace("t.me/", "").strip("/").split("/")[0].lstrip("@")

    if not value or " " in value:
        await update.message.reply_text("❌ Invalid Telegram username.")
        return

    set_setting("support", "@" + value)
    context.user_data.pop("mode", None)

    await update.message.reply_text(
        f"✅ Support set to <code>@{value}</code>.",
        parse_mode=ParseMode.HTML,
        reply_markup=reply_menu(update.effective_user),
    )


# ============================================================
# CALLBACK ROUTER
# ============================================================

async def callbacks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    register_user(query.from_user)

    data = query.data or ""

    if data == "check_join":
        missing = await missing_joined(context.bot, query.from_user.id)
        if missing and not is_admin(query.from_user):
            await query.edit_message_text(
                "🔐 <b>Still required</b>\n\nJoin the channels and check again.",
                parse_mode=ParseMode.HTML,
                reply_markup=join_markup(missing),
            )
        else:
            await query.edit_message_text(
                "✅ <b>JOIN VERIFIED</b>\n\n✨ You can use the bot now.",
                parse_mode=ParseMode.HTML,
                reply_markup=main_inline(query.from_user),
            )

    elif data == "noop":
        return
    elif data == "get_numbers":
        if await ensure_joined(query, context):
            await show_numbers(query)
    elif data.startswith("number:"):
        if await ensure_joined(query, context):
            await number_detail(query, int(data.split(":")[1]))
    elif data.startswith("reserve:"):
        if await ensure_joined(query, context):
            await reserve(query, int(data.split(":")[1]))
    elif data == "wallet":
        if await ensure_joined(query, context):
            await wallet(query)
    elif data == "status":
        if await ensure_joined(query, context):
            await status(query)
    elif data == "support":
        await support_page(query)
    elif data == "admin":
        await admin_page(query)
    elif data == "statistics":
        await statistics(query)
    elif data == "upload":
        await upload_start(query, context)
    elif data == "export":
        await export_numbers(query)
    elif data == "delete":
        await delete_start(query, context)
    elif data == "workers":
        await workers_page(query)
    elif data == "add_panel":
        await add_panel_start(query, context)
    elif data.startswith("panel_toggle:"):
        if is_admin(query.from_user):
            toggle_panel(int(data.split(":")[1]))
            await workers_page(query)
    elif data.startswith("panel_delete:"):
        if is_admin(query.from_user):
            delete_panel(int(data.split(":")[1]))
            await workers_page(query)
    elif data == "join":
        await join_page(query)
    elif data == "join_add":
        await join_add_start(query, context)
    elif data.startswith("join_delete:"):
        if is_admin(query.from_user):
            delete_join(int(data.split(":")[1]))
            await join_page(query)
    elif data == "otp_groups":
        await otp_groups_page(query)
    elif data == "otp_add":
        await otp_add_start(query, context)
    elif data.startswith("otp_toggle:"):
        if is_admin(query.from_user):
            toggle_otp_group(int(data.split(":")[1]))
            await otp_groups_page(query)
    elif data.startswith("otp_delete:"):
        if is_admin(query.from_user):
            delete_otp_group(int(data.split(":")[1]))
            await otp_groups_page(query)
    elif data == "announcement":
        await announcement_start(query, context)
    elif data == "admins":
        await admins_page(query)
    elif data == "admin_add":
        await admin_add_start(query, context)
    elif data.startswith("admin_remove:"):
        if is_owner(query.from_user):
            ids = managed_admin_ids()
            ids.discard(int(data.split(":")[1]))
            save_managed_admin_ids(ids)
            await admins_page(query)
    elif data == "set_support":
        await support_set_start(query, context)
    elif data == "toggle_auto":
        if is_admin(query.from_user):
            set_setting("auto_delete", "off" if setting("auto_delete") == "on" else "on")
            await admin_page(query)
    elif data == "toggle_strict":
        if is_admin(query.from_user):
            set_setting("strict_one", "off" if setting("strict_one") == "on" else "on")
            await admin_page(query)
    elif data == "back":
        await query.edit_message_text(
            "💎 <b>𝗠𝗔𝗜𝗡 𝗠𝗘𝗡𝗨</b>",
            parse_mode=ParseMode.HTML,
            reply_markup=main_inline(query.from_user),
        )


# ============================================================
# TEXT / COMMAND ROUTER
# ============================================================

async def text_handler(update, context):
    user = update.effective_user
    register_user(user)
    mode = context.user_data.get("mode")

    if mode == "upload":
        if is_admin(user):
            await process_text_upload(update, context)
        return
    if mode == "delete":
        if is_admin(user):
            await process_delete(update, context)
        return
    if mode == "panel_setup":
        await process_panel_setup(update, context)
        return
    if mode == "join_add":
        await process_join_add(update, context)
        return
    if mode == "otp_add":
        await process_otp_add(update, context)
        return
    if mode == "announcement":
        await process_announcement(update, context)
        return
    if mode == "admin_add":
        await process_admin_add(update, context)
        return
    if mode == "support_set":
        await process_support_set(update, context)
        return

    if not await ensure_joined(update.message, context):
        return

    text = (update.message.text or "").strip()

    if text == BTN_GET:
        await show_numbers(update.message)
    elif text == BTN_WALLET:
        await wallet(update.message)
    elif text == BTN_STATUS:
        await status(update.message)
    elif text == BTN_ADMIN and is_admin(user):
        await update.message.reply_text(
            "🛡️ <b>𝗔𝗗𝗠𝗜𝗡 𝗗𝗔𝗦𝗛𝗕𝗢𝗔𝗥𝗗</b>",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu(),
        )
    elif text == BTN_SUPPORT and not is_admin(user):
        await support_page(update.message)
    else:
        await update.message.reply_text(
            "✨ Please choose an option.",
            reply_markup=reply_menu(user),
        )


async def admin_command(update, context):
    user = update.effective_user
    register_user(user)
    if not is_admin(user):
        await support_page(update.message)
        return
    await update.message.reply_text(
        "🛡️ <b>𝗔𝗗𝗠𝗜𝗡 𝗗𝗔𝗦𝗛𝗕𝗢𝗔𝗥𝗗</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_menu(),
    )


# ============================================================
# AUTHORIZED PANEL INVENTORY SYNC
# ============================================================

class PanelMonitor:
    def __init__(self, panel):
        self.panel = panel
        self.pw = None
        self.browser = None
        self.page = None

    async def start(self):
        self.pw = await async_playwright().start()
        self.browser = await self.pw.chromium.launch(headless=PANEL_HEADLESS)
        self.page = await self.browser.new_page()
        await self.login()

    async def login(self):
        await self.page.goto(
            self.panel["url"],
            wait_until="domcontentloaded",
            timeout=60000,
        )
        await self.page.fill(
            PANEL_USERNAME_SELECTOR,
            self.panel["username"],
        )
        await self.page.fill(
            PANEL_PASSWORD_SELECTOR,
            self.panel["password"],
        )
        await self.page.click(PANEL_LOGIN_SELECTOR)
        await self.page.wait_for_load_state("domcontentloaded")

    async def read_inventory(self):
        rows = await self.page.locator(
            f"{PANEL_TABLE_SELECTOR} tbody tr"
        ).all()

        result = []
        for row in rows:
            cells = await row.locator("td").all_inner_texts()
            if not cells:
                continue

            number = cells[0].strip()
            if not number:
                continue

            country = cells[1].strip() if len(cells) > 1 else ""
            status = cells[2].strip() if len(cells) > 2 else ""

            result.append({
                "number": number,
                "country": country,
                "status": status,
            })
        return result

    async def sync_once(self):
        inventory = await self.read_inventory()

        for item in inventory:
            add_number(
                item["number"],
                item["country"],
                item["status"],
                source_panel=self.panel["name"],
            )

        sql(
            "UPDATE panels SET last_sync=?, last_error='' WHERE id=?",
            (utc_now(), self.panel["id"]),
        )

    async def run(self):
        try:
            await self.start()

            while True:
                try:
                    await self.sync_once()
                except Exception as exc:
                    error = str(exc)[:500]
                    sql(
                        "UPDATE panels SET last_error=? WHERE id=?",
                        (error, self.panel["id"]),
                    )
                    try:
                        await self.login()
                    except Exception:
                        pass

                await asyncio.sleep(CHECK_SECONDS)

        except asyncio.CancelledError:
            raise
        except Exception as exc:
            sql(
                "UPDATE panels SET last_error=? WHERE id=?",
                (str(exc)[:500], self.panel["id"]),
            )
        finally:
            if self.browser:
                await self.browser.close()
            if self.pw:
                await self.pw.stop()


async def panel_manager():
    running = {}

    while True:
        configured = panels()
        active = {row["id"] for row in configured if row["active"]}

        for row in configured:
            if row["active"] and (
                row["id"] not in running or running[row["id"]].done()
            ):
                running[row["id"]] = asyncio.create_task(
                    PanelMonitor(row).run()
                )

        for panel_id in list(running):
            if panel_id not in active:
                running[panel_id].cancel()
                running.pop(panel_id, None)

        await asyncio.sleep(20)


async def post_init(app):
    app.create_task(panel_manager())


# ============================================================
# MAIN
# ============================================================

def main():
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN is missing in .env")

    init_db()

    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .post_init(post_init)
        .build()
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler(["admin", "panel"], admin_command))
    app.add_handler(CallbackQueryHandler(callbacks))
    app.add_handler(MessageHandler(filters.Document.ALL, document_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))

    print("💎 ZENOX NUMBER BOT STARTED")
    app.run_polling()


if __name__ == "__main__":
    main()
